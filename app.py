import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from ortools.sat.python import cp_model
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
import io

st.set_page_config(page_title="Provider Utilization and Logistics Scheduling Engine", layout="wide")
st.title("Provider Utilization and Logistics Scheduling Engine")

# ==========================================
# UI: DATA INPUTS
# ==========================================
st.sidebar.header("1. Setup")
anchor_date = st.sidebar.date_input("Select SUNDAY 1a Start Date")

st.sidebar.header("2. Uploads")
roster_file = st.sidebar.file_uploader("Upload Prior Schedule (Roster)", type=['xlsx'])
leave_file = st.sidebar.file_uploader("Upload Leave Requests", type=['xlsx'])

# ==========================================
# MODULE: A.R.M. (Automated Re-scheduling Module)
# ==========================================
st.divider()
st.title("A.R.M. - Automated Re-scheduling Module")
st.markdown("Upload your AMION schedule to automatically apply shift rules.")

# Initialize ARM Session State
if 'arm_providers' not in st.session_state:
    st.session_state.arm_providers = [
        {"name": "ADAM, G", "initials": "GA", "role": "MD"},
        {"name": "ARCEDO, V", "initials": "VCA", "role": "MD"},
        {"name": "BACASNOT, J", "initials": "JVB", "role": "MD"},
        {"name": "BASTERRECHEA, H", "initials": "HRB", "role": "MD"},
        {"name": "BOU, C", "initials": "CB", "role": "MD"},
        {"name": "CABATU, A", "initials": "ACC", "role": "MD"},
        {"name": "DE LEON, M", "initials": "MDT", "role": "MD"},
        {"name": "DOMINGUEZ, G", "initials": "GJDP", "role": "MD"},
        {"name": "DOTSON, T", "initials": "TMD", "role": "MD"},
        {"name": "GONZALEZ-TORRES, J", "initials": "JGT", "role": "MD"},
        {"name": "GRAZIANO, J", "initials": "JDG", "role": "MD"},
        {"name": "JOHNSON, C", "initials": "CJ", "role": "MD"},
        {"name": "MOHAR, C", "initials": "CMM", "role": "MD"},
        {"name": "PAULIS, J", "initials": "JP", "role": "MD"},
        {"name": "PEREZ, K", "initials": "KFP", "role": "MD"},
        {"name": "POOLE, R", "initials": "RP", "role": "MD"},
        {"name": "RUIZ-QUINONES, J", "initials": "JIR", "role": "MD"},
        {"name": "SANDERS, B", "initials": "BS", "role": "MD"},
        {"name": "SORGE, R", "initials": "RMS", "role": "MD"},
        {"name": "SOSA, E", "initials": "ES", "role": "MD"},
        {"name": "VIJAPURA, C", "initials": "CAV", "role": "MD"},
        {"name": "ALEXANDER, B", "initials": "BRA", "role": "APP"},
        {"name": "BENJAMIN, J", "initials": "JPB", "role": "APP"},
        {"name": "BRINSON, T", "initials": "WTB", "role": "APP"},
        {"name": "CALIFANO, J", "initials": "JNC", "role": "APP"},
        {"name": "DECHON, B", "initials": "RLD", "role": "APP"},
        {"name": "FREDERIQUE-BELL, P", "initials": "PFB", "role": "APP"},
        {"name": "GRIMALDI, T", "initials": "TAG", "role": "APP"},
        {"name": "PIERSON, K", "initials": "KRP", "role": "APP"}
    ]
if 'arm_step' not in st.session_state: st.session_state.arm_step = 'upload'
if 'arm_missing_queue' not in st.session_state: st.session_state.arm_missing_queue = []
if 'arm_raw_entries' not in st.session_state: st.session_state.arm_raw_entries = []
if 'arm_parsed_dates' not in st.session_state: st.session_state.arm_parsed_dates = []

def apply_arm_rules(shift_str, is_weekend, initial):
    import re
    if pd.isna(shift_str) or str(shift_str).strip() == "": return ""
    match = re.search(r'(\d{1,2})([ap])m?\s*-\s*(\d{1,2})([ap])m?', str(shift_str), re.IGNORECASE)
    if not match: return "" 
    
    sH, sM, eH, eM = int(match.group(1)), match.group(2).lower(), int(match.group(3)), match.group(4).lower()

    if (sH == 7 or sH == 9) and sM == 'p':
        return "8p-6a" if is_weekend else "8p-8a"

    s24 = (12 if sM == 'p' else 0) if sH == 12 else sH + (12 if sM == 'p' else 0)
    e24 = (12 if eM == 'p' else 0) if eH == 12 else eH + (12 if eM == 'p' else 0)
    dur = e24 - s24 if e24 >= s24 else (e24 + 24) - s24

    target_dur = dur
    allowed_12hr = ["JNC", "JIR", "ACC", "VCA", "WTB", "MDT", "CJ", "RMS", "RP"]

    if is_weekend and dur >= 12: target_dur = 10
    elif not is_weekend and dur >= 12:
        if str(initial).strip().upper() not in allowed_12hr: target_dur = 10

    if target_dur != dur:
        new_e24 = (s24 + target_dur) % 24
        eH = new_e24 % 12 or 12
        eM = 'p' if new_e24 >= 12 else 'a'

    return f"{sH}{sM}-{eH}{eM}"

# UI STEP 1: UPLOAD
if st.session_state.arm_step == 'upload':
    arm_file = st.file_uploader("Upload Schedule for A.R.M. (Excel)", type=['xlsx', 'xls', 'csv'], key="arm_upload")
    if arm_file and st.button("Parse Schedule"):
        df_arm = pd.read_excel(arm_file, header=None)
        date_row_idx = -1
        for idx, row in df_arm.head(10).iterrows():
            if sum(pd.to_datetime(row[1:], errors='coerce').notna()) >= 3:
                date_row_idx = idx
                break
                
        if date_row_idx != -1:
            date_row = df_arm.iloc[date_row_idx]
            all_dates = pd.to_datetime(date_row[1:], errors='coerce')
            
            # SLICE LAST 15 COLUMNS ONLY
            valid_date_cols = all_dates.dropna().index[-15:]
            st.session_state.arm_parsed_dates = all_dates[valid_date_cols]
            
            raw_entries = []
            unique_initials = set()
            known_initials = [p['initials'].upper() for p in st.session_state.arm_providers]
            
            for idx in range(date_row_idx + 1, len(df_arm)):
                row = df_arm.iloc[idx]
                shift_str = str(row[0]).strip()
                if not shift_str or shift_str == "nan": continue
                
                # Iterate only through the last 15 valid columns
                for c_idx, date_val in st.session_state.arm_parsed_dates.items():
                    cell_val = str(row[c_idx]).strip()
                    if cell_val and cell_val != "nan":
                        initials_list = [v.strip().upper() for v in cell_val.split(',') if v.strip()]
                        for init in initials_list:
                            if init.isalpha():
                                unique_initials.add(init)
                                raw_entries.append({
                                    'initial': init, 'col_idx': c_idx, 'date': date_val,
                                    'raw_shift': shift_str, 'is_weekend': date_val.weekday() >= 5
                                })
                                
            st.session_state.arm_raw_entries = raw_entries
            st.session_state.arm_missing_queue = [init for init in unique_initials if init not in known_initials]
            st.session_state.arm_step = 'resolve' if st.session_state.arm_missing_queue else 'generate'
            st.rerun()
        else:
            st.error("Could not locate date header row.")

# UI STEP 2: RESOLVE UNKNOWN PROVIDERS
elif st.session_state.arm_step == 'resolve':
    current_init = st.session_state.arm_missing_queue[0]
    st.warning(f"⚠️ **New Initials Detected:** `{current_init}`")
    st.info("These initials were found in the schedule but are not in the database.")
    
    with st.form("arm_resolve_form"):
        new_name = st.text_input("Name (Last, First Initial)", placeholder="e.g. SMITH, J")
        new_role = st.selectbox("Role", ["MD", "APP"])
        
        col1, col2 = st.columns(2)
        save_btn = col1.form_submit_button("Save Provider", type="primary")
        skip_btn = col2.form_submit_button("Skip Provider")
        
        if save_btn:
            if new_name:
                st.session_state.arm_providers.append({"name": new_name.upper(), "initials": current_init, "role": new_role})
                st.session_state.arm_missing_queue.pop(0)
                if not st.session_state.arm_missing_queue:
                    st.session_state.arm_step = 'generate'
                st.rerun()
            else:
                st.error("Name is required to save.")
        elif skip_btn:
            st.session_state.arm_raw_entries = [e for e in st.session_state.arm_raw_entries if e['initial'] != current_init]
            st.session_state.arm_missing_queue.pop(0)
            if not st.session_state.arm_missing_queue:
                st.session_state.arm_step = 'generate'
            st.rerun()

# UI STEP 3: GENERATE
elif st.session_state.arm_step == 'generate':
    st.success("Schedule processed and ready for export.")
    
    # Add Provider Manually Expander
    with st.expander("➕ Add Provider Manually (If Missing)"):
        with st.form("manual_add_form"):
            col1, col2, col3 = st.columns(3)
            with col1: m_name = st.text_input("Name (Last, F)")
            with col2: m_init = st.text_input("Initials").upper()
            with col3: m_role = st.selectbox("Role", ["MD", "APP"])
            if st.form_submit_button("Add to Schedule"):
                if m_name and m_init:
                    st.session_state.arm_providers.append({"name": m_name.upper(), "initials": m_init, "role": m_role})
                    st.rerun()
                else:
                    st.error("Name and Initials are required.")

    # Initialize output structure
    output_dict = {}
    for p in st.session_state.arm_providers:
        output_dict[p['initials']] = {
            'Provider': p['name'], 'Initials': p['initials'], 'Role': p['role']
        }
        for d in st.session_state.arm_parsed_dates.dropna():
            output_dict[p['initials']][d.strftime('%a %m/%d')] = ""

    # Map shifts
    for entry in st.session_state.arm_raw_entries:
        init = entry['initial']
        if init in output_dict:
            clean_shift = apply_arm_rules(entry['raw_shift'], entry['is_weekend'], init)
            if clean_shift:
                date_key = entry['date'].strftime('%a %m/%d')
                existing = output_dict[init][date_key]
                output_dict[init][date_key] = f"{existing}, {clean_shift}" if existing else clean_shift

    df_out = pd.DataFrame(list(output_dict.values()))
    
    # Sort: MD first, then APP, then alphabetical by name
    df_out['Role_Sort'] = df_out['Role'].apply(lambda x: 0 if x == 'MD' else 1)
    df_out = df_out.sort_values(['Role_Sort', 'Provider']).drop('Role_Sort', axis=1)

    st.dataframe(df_out)
    
    out_buffer = io.BytesIO()
    df_out.to_excel(out_buffer, index=False, sheet_name="Processed Schedule")
    
    col1, col2 = st.columns([1, 4])
    with col1:
        st.download_button("Export Excel", data=out_buffer.getvalue(), file_name="ARM_Processed.xlsx", mime="application/vnd.ms-excel")
    with col2:
        if st.button("Start Over"):
            st.session_state.arm_step = 'upload'
            st.rerun()

# ==========================================
# MODULE: PLANS (Pay-period Leave Allocation)
# ==========================================
st.divider()
st.title("PLANS - Leave Allocation Tracker")
st.markdown("Upload your AMION roster to extract and calculate leave blocks.")

if 'plans_initials' not in st.session_state:
    st.session_state.plans_initials = {
        "ADAM, G": "GA", "ARCEDO, V": "VCA", "BACASNOT, J": "JVB", "BASTERRECHEA, H": "HRB",
        "BOU, C": "CB", "CABATU, A": "ACC", "DE LEON, M": "MDT", "DOMINGUEZ, G": "GJDP",
        "DOTSON, T": "TMD", "GONZALEZ-TORRES, J": "JGT", "GRAZIANO, J": "JDG", "JOHNSON, C": "CJ",
        "MOHAR, C": "CMM", "PAULIS, J": "JP", "PEREZ, K": "KFP", "POOLE, R": "RP",
        "RUIZ-QUINONES, J": "JIR", "SANDERS, B": "BS", "SORGE, R": "RMS", "SOSA, E": "ES",
        "VIJAPURA, C": "CAV", "ALEXANDER, B": "BRA", "BENJAMIN, J": "JPB", "BRINSON, T": "WTB",
        "CALIFANO, J": "JNC", "DECHON, B": "RLD", "FREDERIQUE-BELL, P": "PFB", "GRIMALDI, T": "TAG",
        "PIERSON, K": "KRP"
    }
if 'plans_step' not in st.session_state: st.session_state.plans_step = 'upload'
if 'plans_unknowns' not in st.session_state: st.session_state.plans_unknowns = []
if 'plans_df' not in st.session_state: st.session_state.plans_df = None
if 'plans_dates' not in st.session_state: st.session_state.plans_dates = None

# UI STEP 1: UPLOAD
if st.session_state.plans_step == 'upload':
    plans_file = st.file_uploader("Upload Roster for PLANS (Excel)", type=['xlsx', 'xls', 'csv'], key="plans_upload")
    if plans_file and st.button("Parse Roster"):
        df_plans = pd.read_excel(plans_file, header=None)
        date_row_idx = -1
        for idx, row in df_plans.head(10).iterrows():
            if sum(pd.to_datetime(row[1:], errors='coerce').notna()) >= 3:
                date_row_idx = idx
                break

        if date_row_idx != -1:
            st.session_state.plans_dates = pd.to_datetime(df_plans.iloc[date_row_idx][1:], errors='coerce')
            st.session_state.plans_df = df_plans.iloc[date_row_idx + 1:]
            
            # Find unknowns
            unknowns = []
            for idx, row in st.session_state.plans_df.iterrows():
                name = str(row[0]).replace('"', '').replace("'", "").strip()
                if name and name != "nan" and name not in st.session_state.plans_initials:
                    if name not in unknowns: unknowns.append(name)
            
            st.session_state.plans_unknowns = unknowns
            st.session_state.plans_step = 'resolve' if unknowns else 'generate'
            st.rerun()
        else:
            st.error("Could not locate date header row.")

# UI STEP 2: RESOLVE UNKNOWNS
elif st.session_state.plans_step == 'resolve':
    st.warning(f"⚠️ **Found {len(st.session_state.plans_unknowns)} new provider(s)**.")
    st.info("Assign their initials below, or check 'Ignore' to skip their records entirely.")
    
    with st.form("plans_resolve_form"):
        mappings = {}
        for name in st.session_state.plans_unknowns:
            col1, col2, col3 = st.columns([3, 1, 2])
            with col1:
                st.markdown(f"**{name}**")
            with col2:
                ignore = st.checkbox("Ignore", key=f"ign_{name}")
            with col3:
                init = st.text_input("Initials", key=f"init_{name}", disabled=ignore, label_visibility="collapsed", placeholder="Initials")
            
            mappings[name] = {"ignore": ignore, "initials": init}
            
        if st.form_submit_button("Save & Process", type="primary"):
            for name, data in mappings.items():
                if data["ignore"]:
                    st.session_state.plans_initials[name] = "IGNORE"
                else:
                    val_clean = data["initials"].strip().upper()
                    if val_clean: st.session_state.plans_initials[name] = val_clean
            st.session_state.plans_step = 'generate'
            st.rerun()

# UI STEP 3: GENERATE
elif st.session_state.plans_step == 'generate':
    special_providers = ['JNC', 'JIR', 'ACC', 'VCA', 'WTB', 'MDT', 'CJ', 'RMS', 'RP']
    leave_records = []
    
    for idx, row in st.session_state.plans_df.iterrows():
        raw_name = str(row[0]).replace('"', '').replace("'", "").strip()
        if not raw_name or raw_name == "nan": continue
        
        initials = st.session_state.plans_initials.get(raw_name)
        if not initials or initials == "IGNORE": continue
        
        provider_leaves = []
        for c_idx, date_val in st.session_state.plans_dates.items():
            if pd.notna(date_val):
                val = str(row[c_idx]).strip().upper()
                if val in ['VAC', 'ADM', 'SICK', 'X', 'HOLIDAY']:
                    provider_leaves.append({'date': date_val, 'type': val})
        
        if not provider_leaves: continue
        provider_leaves.sort(key=lambda x: x['date'])
        
        paid_leave_count = 0
        for d in provider_leaves:
            if d['type'] in ['VAC', 'ADM', 'SICK']:
                paid_leave_count += 1
                d['hours'] = (10 if paid_leave_count <= 2 else 12) if initials in special_providers else 10
            elif d['type'] in ['X', 'HOLIDAY']:
                d['hours'] = 0

        # Group contiguous blocks
        current_block = [provider_leaves[0]]
        for i in range(1, len(provider_leaves)):
            if (provider_leaves[i]['date'] - current_block[-1]['date']).days == 1:
                current_block.append(provider_leaves[i])
            else:
                leave_records.append({
                    'Initials': initials,
                    'Start_Date': current_block[0]['date'].strftime('%m/%d/%Y'),
                    'End_Date': current_block[-1]['date'].strftime('%m/%d/%Y'),
                    'Leave_Hours': sum(item['hours'] for item in current_block)
                })
                current_block = [provider_leaves[i]]
        
        leave_records.append({
            'Initials': initials,
            'Start_Date': current_block[0]['date'].strftime('%m/%d/%Y'),
            'End_Date': current_block[-1]['date'].strftime('%m/%d/%Y'),
            'Leave_Hours': sum(item['hours'] for item in current_block)
        })

    df_out_plans = pd.DataFrame(leave_records)
    st.success(f"Successfully processed {len(leave_records)} leave blocks!")
    st.dataframe(df_out_plans)
    
    out_buffer_plans = io.BytesIO()
    df_out_plans.to_excel(out_buffer_plans, index=False, sheet_name="Leave_Data")
    
    col1, col2 = st.columns([1, 4])
    with col1:
        st.download_button("Export Excel", data=out_buffer_plans.getvalue(), file_name="PLANS_EXPORT.xlsx", mime="application/vnd.ms-excel")
    with col2:
        if st.button("Start Over", key="plans_reset"):
            st.session_state.plans_step = 'upload'
            st.rerun()

st.divider()
st.title("P.U.L.S.E.")

# ==========================================
# ENGINE BLOCKS
# ==========================================
def run_block_1(uploaded_roster):
    df_prior = pd.read_excel(uploaded_roster)
    expected_cols = ['Provider', 'Initials', 'Role']
    missing_cols = [col for col in expected_cols if col not in df_prior.columns]
    
    if missing_cols:
        st.error(f"Missing core columns in roster: {missing_cols}")
        st.stop()

    providers = df_prior['Initials'].dropna().astype(str).tolist()
    roles = df_prior.set_index('Initials')['Role'].dropna().to_dict()

    missing_fee = [fee for fee in ['FEE 1', 'FEE 2'] if fee not in providers]
    if missing_fee:
        new_rows = []
        for fee in missing_fee:
            providers.append(fee)
            roles[fee] = 'MD'
            new_row = {col: '' for col in df_prior.columns}
            new_row['Initials'] = fee
            new_row['Role'] = 'MD'
            new_row['Provider'] = f'{fee} (Auto-Added System Net)'
            new_rows.append(new_row)
        df_prior = pd.concat([df_prior, pd.DataFrame(new_rows)], ignore_index=True)
        
    return df_prior, providers, roles

def run_block_2(uploaded_leave, valid_providers):
    if uploaded_leave is None:
        return {}
        
    df_leave = pd.read_excel(uploaded_leave)
    df_leave = df_leave.dropna(how='all')
    df_leave['Start_Date'] = pd.to_datetime(df_leave['Start_Date'], errors='coerce')
    df_leave['End_Date'] = pd.to_datetime(df_leave['End_Date'], errors='coerce')
    df_leave['Leave_Hours'] = pd.to_numeric(df_leave['Leave_Hours'], errors='coerce').fillna(0)

    df_leave = df_leave[df_leave['Initials'].isin(valid_providers)].copy()

    leave_dict = {}
    for _, row in df_leave.iterrows():
        init = row['Initials']
        if init not in leave_dict:
            leave_dict[init] = []
        leave_dict[init].append({
            'start': row['Start_Date'],
            'end': row['End_Date'],
            'hours': row['Leave_Hours']
        })
    return leave_dict

def run_block_3_and_4(anchor_date_py, providers, roles, df_prior, leave_dict):
    model = cp_model.CpModel()
    num_days = 56
    date_list = [anchor_date_py + timedelta(days=i) for i in range(num_days)]

    holidays = [
        datetime(2026, 5, 25), datetime(2026, 6, 19), datetime(2026, 7, 3),
        datetime(2026, 9, 7), datetime(2026, 10, 12), datetime(2026, 11, 11),
        datetime(2026, 11, 26), datetime(2026, 12, 25),
        datetime(2027, 1, 1), datetime(2027, 1, 18), datetime(2027, 2, 15),
        datetime(2027, 5, 31), datetime(2027, 6, 18), datetime(2027, 7, 5),
        datetime(2027, 9, 6), datetime(2027, 10, 11), datetime(2027, 11, 11),
        datetime(2027, 11, 25), datetime(2027, 12, 24)
    ]

    shifts_4x10 = {
        '5a-3p': (5, 10), '6a-4p': (6, 10), '7a-5p': (7, 10), '8a-6p': (8, 10),
        '9a-7p': (9, 10), '10a-8p': (10, 10), '12p-10p': (12, 10),
        '2p-12a': (14, 10), '4p-2a': (16, 10), '8p-6a': (20, 10)
    }

    shifts_7x7 = {
        '5a-5p': (5, 12), '7a-7p': (7, 12), '9a-9p': (9, 12), '8p-8a': (20, 12),
        '6p-6a': (18, 12), '6p-4a': (18, 10), '7a-5p': (7, 10), '9a-7p': (9, 10), 
        '8p-6a': (20, 10), '6a-4p': (6, 10)
    }
    all_shifts = {**shifts_4x10, **shifts_7x7}
    shift_names = list(all_shifts.keys())

    group_a = ['JNC', 'JIR', 'ACC', 'VCA']
    group_b = ['WTB', 'MDT', 'CJ', 'RMS', 'RP']
    group_a_clean = [x.strip().upper() for x in group_a]
    group_b_clean = [x.strip().upper() for x in group_b]

    seven_on_providers = [p for p in providers if p.strip().upper() in group_a_clean or p.strip().upper() in group_b_clean]
    four_ten_providers = [p for p in providers if p not in seven_on_providers and "FEE" not in p.upper()]
    fee_providers = [p for p in providers if "FEE" in p.upper()]

    schedule_vars = {}
    works_day = {}
    penalty_vars = []
    
    for p in providers:
        for d in range(num_days):
            works_day[(p, d)] = model.NewBoolVar(f'works_{p}_{d}')
            for s in shift_names:
                schedule_vars[(p, d, s)] = model.NewBoolVar(f'shift_{p}_{d}_{s}')
            model.Add(works_day[(p, d)] == sum(schedule_vars[(p, d, s)] for s in shift_names))

    # --- RULE SET 0: THE DAY 0 HARD LOCK (CONTINUITY) ---
    if df_prior is not None and not df_prior.empty:
        day_0_col = df_prior.columns[-1] 
        for p in providers:
            p_data = df_prior[df_prior['Initials'].astype(str).str.strip().str.upper() == p.strip().upper()]
            if not p_data.empty:
                val = str(p_data[day_0_col].values[0]).strip()
                if val in shift_names:
                    model.Add(schedule_vars[(p, 0, val)] == 1)
                else:
                    model.Add(works_day[(p, 0)] == 0)

    for p in providers:
        for d in range(num_days):
            if d == 0: continue 
            if any(l['start'] <= date_list[d] <= l['end'] for l in leave_dict.get(p, [])):
                model.Add(works_day[(p, d)] == 0)

    track_vars = {}

    for p in seven_on_providers:
        p_clean = p.strip().upper()
        is_group_a = p_clean in group_a_clean
        valid_work_days = []
        
        for c in range(4):
            period_start_d = c * 14
            period_end_d = period_start_d + 13
            if period_end_d >= num_days: period_end_d = num_days - 1
            
            has_period_leave = any(
                l['start'] <= date_list[period_end_d] and l['end'] >= date_list[period_start_d]
                for l in leave_dict.get(p, [])
            )
            
            t_5a = model.NewBoolVar(f't_5a_{p}_{c}')
            t_7a = model.NewBoolVar(f't_7a_{p}_{c}')
            t_9a = model.NewBoolVar(f't_9a_{p}_{c}')
            t_8p = model.NewBoolVar(f't_8p_{p}_{c}')
            t_6p = model.NewBoolVar(f't_6p_{p}_{c}')
            
            model.AddExactlyOne([t_5a, t_7a, t_9a, t_8p, t_6p])
            
            track_vars[(p_clean, c, 'night')] = t_8p + t_6p
            track_vars[(p_clean, c, 'day')] = t_5a + t_7a + t_9a
            
            if p_clean in ['JNC', 'WTB']: model.Add(t_6p == 1)
            elif p_clean in ['JIR', 'MDT']: model.Add(t_8p == 1)
            elif p_clean in ['ACC', 'VCA', 'RMS', 'RP']: model.Add(t_6p == 0)
            
            if is_group_a:
                b_days = [c * 14 + i for i in range(1, 8)]
            else:
                b_days = [c * 14 + i for i in range(8, 15)]
                
            for d in b_days:
                if d >= num_days: continue
                valid_work_days.append(d)
                
                is_hard_blocked = any(l['start'] <= date_list[d] <= l['end'] for l in leave_dict.get(p, []))
                is_leave_tomorrow = (d + 1 < num_days) and any(l['start'] <= date_list[d+1] <= l['end'] for l in leave_dict.get(p, []))
                
                if is_hard_blocked:
                    model.Add(works_day[(p, d)] == 0)
                else:
                    if is_leave_tomorrow:
                        model.Add(works_day[(p, d)] == 0).OnlyEnforceIf(t_8p)
                        model.Add(works_day[(p, d)] == 0).OnlyEnforceIf(t_6p)
                        model.Add(works_day[(p, d)] == 1).OnlyEnforceIf(t_5a)
                        model.Add(works_day[(p, d)] == 1).OnlyEnforceIf(t_7a)
                        model.Add(works_day[(p, d)] == 1).OnlyEnforceIf(t_9a)
                    else:
                        model.Add(works_day[(p, d)] == 1)

                    is_weekend = date_list[d].weekday() >= 5
                    
                    if is_weekend:
                        if not has_period_leave:
                            model.Add(schedule_vars[(p, d, '6a-4p')] == works_day[(p, d)]).OnlyEnforceIf(t_5a)
                            model.Add(schedule_vars[(p, d, '7a-5p')] == works_day[(p, d)]).OnlyEnforceIf(t_7a)
                            model.Add(schedule_vars[(p, d, '9a-7p')] == works_day[(p, d)]).OnlyEnforceIf(t_9a)
                            model.Add(schedule_vars[(p, d, '8p-6a')] == works_day[(p, d)]).OnlyEnforceIf(t_8p)
                            model.Add(schedule_vars[(p, d, '6p-4a')] == works_day[(p, d)]).OnlyEnforceIf(t_6p)
                        else:
                            model.Add(schedule_vars[(p, d, '5a-5p')] + schedule_vars[(p, d, '6a-4p')] == works_day[(p, d)]).OnlyEnforceIf(t_5a)
                            model.Add(schedule_vars[(p, d, '7a-7p')] + schedule_vars[(p, d, '7a-5p')] == works_day[(p, d)]).OnlyEnforceIf(t_7a)
                            model.Add(schedule_vars[(p, d, '9a-9p')] + schedule_vars[(p, d, '9a-7p')] == works_day[(p, d)]).OnlyEnforceIf(t_9a)
                            model.Add(schedule_vars[(p, d, '8p-8a')] + schedule_vars[(p, d, '8p-6a')] == works_day[(p, d)]).OnlyEnforceIf(t_8p)
                            model.Add(schedule_vars[(p, d, '6p-6a')] + schedule_vars[(p, d, '6p-4a')] == works_day[(p, d)]).OnlyEnforceIf(t_6p)
                    else:
                        model.Add(schedule_vars[(p, d, '5a-5p')] == works_day[(p, d)]).OnlyEnforceIf(t_5a)
                        model.Add(schedule_vars[(p, d, '7a-7p')] == works_day[(p, d)]).OnlyEnforceIf(t_7a)
                        model.Add(schedule_vars[(p, d, '9a-9p')] == works_day[(p, d)]).OnlyEnforceIf(t_9a)
                        model.Add(schedule_vars[(p, d, '8p-8a')] == works_day[(p, d)]).OnlyEnforceIf(t_8p)
                        model.Add(schedule_vars[(p, d, '6p-6a')] == works_day[(p, d)]).OnlyEnforceIf(t_6p)
                        
        for d in range(num_days):
            if d not in valid_work_days and d != 0: 
                model.Add(works_day[(p, d)] == 0)

    for pair_1, pair_2 in [('ACC', 'VCA'), ('RP', 'RMS')]:
        if pair_1 in group_a_clean or pair_1 in group_b_clean:
            p1_was_night = False
            
            if df_prior is not None:
                try:
                    p1_data = df_prior[df_prior.iloc[:, 0].astype(str).str.strip().str.upper() == pair_1]
                    p2_data = df_prior[df_prior.iloc[:, 0].astype(str).str.strip().str.upper() == pair_2]
                    
                    p1_n_count, p2_n_count = 0, 0
                    if not p1_data.empty:
                        r_p1 = " ".join(p1_data.iloc[0, -14:].fillna("").astype(str))
                        p1_n_count = r_p1.count('8p') + r_p1.count('6p')
                    if not p2_data.empty:
                        r_p2 = " ".join(p2_data.iloc[0, -14:].fillna("").astype(str))
                        p2_n_count = r_p2.count('8p') + r_p2.count('6p')
                        
                    if p1_n_count > p2_n_count: p1_was_night = True
                    elif p2_n_count > p1_n_count: p1_was_night = False
                    else: p1_was_night = True 
                except: pass

            p1_roster = next((p for p in providers if p.strip().upper() == pair_1), None)
            p2_roster = next((p for p in providers if p.strip().upper() == pair_2), None)
            is_group_a = pair_1 in group_a_clean

            for c in range(4):
                p1_night = track_vars[(pair_1, c, 'night')]
                p2_night = track_vars[(pair_2, c, 'night')]
                
                model.Add(p1_night + p2_night == 1)
                
                if is_group_a:
                    b_days = [c * 14 + i for i in range(1, 8)]
                else:
                    b_days = [c * 14 + i for i in range(8, 15)]
                
                b_days = [d for d in b_days if d < num_days]
                
                p1_has_leave = any(any(l['start'] <= date_list[d] <= l['end'] for l in leave_dict.get(p1_roster, [])) for d in b_days)
                p2_has_leave = any(any(l['start'] <= date_list[d] <= l['end'] for l in leave_dict.get(p2_roster, [])) for d in b_days)
                
                if p1_has_leave and not p2_has_leave:
                    model.Add(p1_night == 0)
                elif p2_has_leave and not p1_has_leave:
                    model.Add(p1_night == 1)
                else:
                    if c > 0:
                        model.Add(p1_night != track_vars[(pair_1, c-1, 'night')]) 
                    else:
                        if p1_was_night:
                            model.Add(p1_night == 0) 
                        else:
                            model.Add(p1_night == 1)

    allowed_7x7_shifts = ['5a-5p', '7a-7p', '9a-9p', '8p-8a', '6p-6a', '6a-4p', '7a-5p', '9a-7p', '8p-6a', '6p-4a']
    
    for d in range(num_days):
        if d == 0: continue 
        is_friday = date_list[d].weekday() == 4
        is_weekend = date_list[d].weekday() >= 5
        is_holiday = date_list[d] in holidays
        
        for p in providers:
            p_clean = p.strip().upper()
            is_leave_tomorrow = (d + 1 < num_days) and any(l['start'] <= date_list[d+1] <= l['end'] for l in leave_dict.get(p, []))

            if p in four_ten_providers:
                for s in shift_names:
                    if s not in shifts_4x10:
                        model.Add(schedule_vars[(p, d, s)] == 0)
                
                if not is_weekend and not is_holiday and roles.get(p) != 'MD':
                    model.Add(schedule_vars[(p, d, '9a-7p')] == 0)

            if p in seven_on_providers:
                for s in shift_names:
                    if s not in allowed_7x7_shifts:
                        model.Add(schedule_vars[(p, d, s)] == 0)

            for s in shift_names:
                start_hour = all_shifts[s][0]
                end_hour = start_hour + all_shifts[s][1]

                if is_leave_tomorrow and end_hour > 22:
                    model.Add(schedule_vars[(p, d, s)] == 0)

                if p_clean == 'CB' and start_hour >= 12:
                    model.Add(schedule_vars[(p, d, s)] == 0)
                if p_clean == 'GA' and start_hour > 12:
                    model.Add(schedule_vars[(p, d, s)] == 0)
                if p_clean == 'CJ' and start_hour > 9:
                    model.Add(schedule_vars[(p, d, s)] == 0)
                if is_friday and p_clean in ['JDT', 'JGT'] and start_hour > 7:
                    model.Add(schedule_vars[(p, d, s)] == 0)
                if p_clean == 'RLD' and s != '8p-6a':
                    model.Add(schedule_vars[(p, d, s)] == 0)

    for p in four_ten_providers:
        p_clean = p.strip().upper()
        max_consecutive = 5 if p_clean == 'TAG' else 4
        window_size = max_consecutive + 1
        
        for d in range(1, num_days - 1):
            l_yest = any(l['start'] <= date_list[d-1] <= l['end'] for l in leave_dict.get(p, []))
            l_tod = any(l['start'] <= date_list[d] <= l['end'] for l in leave_dict.get(p, []))
            l_tom = any(l['start'] <= date_list[d+1] <= l['end'] for l in leave_dict.get(p, []))

            if not (l_yest or l_tod or l_tom):
                pen_1on = model.NewBoolVar(f'pen_1on_{p}_{d}')
                model.Add(works_day[(p, d)] - works_day[(p, d-1)] - works_day[(p, d+1)] <= pen_1on)
                penalty_vars.append(pen_1on * 5000)

                pen_1off = model.NewBoolVar(f'pen_1off_{p}_{d}')
                model.Add(works_day[(p, d-1)] + works_day[(p, d+1)] - works_day[(p, d)] - 1 <= pen_1off)
                penalty_vars.append(pen_1off * 5000)

        for d in range(num_days - max_consecutive):
            pen_max_on = model.NewIntVar(0, window_size, f'pen_max_on_{p}_{d}')
            model.Add(sum(works_day[(p, i)] for i in range(d, d + window_size)) - max_consecutive <= pen_max_on)
            penalty_vars.append(pen_max_on * 5000)

        for d in range(num_days - 5):
            leave_in_window = any(any(l['start'] <= date_list[i] <= l['end'] for l in leave_dict.get(p, [])) for i in range(d, d+6))
            if not leave_in_window:
                pen_max_off = model.NewBoolVar(f'pen_max_off_{p}_{d}')
                model.Add(1 - sum(works_day[(p, i)] for i in range(d, d+6)) <= pen_max_off)
                penalty_vars.append(pen_max_off * 500)

    for p in providers:
        for d in range(num_days - 1):
            for s1 in shift_names:
                for s2 in shift_names:
                    if s1 != s2:
                        diff_penalty = model.NewBoolVar(f'diff_{p}_{d}_{s1}_{s2}')
                        model.Add(diff_penalty >= schedule_vars[(p, d, s1)] + schedule_vars[(p, d+1, s2)] - 1)
                        penalty_vars.append(diff_penalty * 50)

                    t1, L1 = all_shifts[s1][0], all_shifts[s1][1]
                    t2 = all_shifts[s2][0]
                    if (t2 + 24) - (t1 + L1) < 10 or (t2 > t1 + 2 and t2 < 24) or (t1 >= 18 and t2 < t1):
                        model.AddImplication(schedule_vars[(p, d, s1)], schedule_vars[(p, d+1, s2)].Not())

    for d in range(num_days):
        if d == 0: continue 
        is_weekend = date_list[d].weekday() >= 5
        is_holiday = date_list[d] in holidays
        target_floor = 90 if (is_weekend or is_holiday) else 120

        daily_physician_hours = sum(schedule_vars[(p, d, s)] * all_shifts[s][1] for p in providers if p not in fee_providers for s in shift_names)
        floor_gap = model.NewIntVar(0, target_floor, f'floor_gap_{d}')
        model.Add(daily_physician_hours + floor_gap >= target_floor)
        penalty_vars.append(floor_gap * 100)

    for p in four_ten_providers:
        for period in range(4):
            start_d, end_d = period * 14, (period * 14) + 14
            hrs = [schedule_vars[(p, d, s)] * all_shifts[s][1] for d in range(start_d, end_d) for s in shift_names]
            l_hrs = sum(l['hours'] for l in leave_dict.get(p, []) if l['start'] <= date_list[end_d-1] and l['end'] >= date_list[start_d])
            target = 70 if any(date_list[d] in holidays for d in range(start_d, end_d)) else 80
            
            gap = model.NewIntVar(-1000, 1000, f'gap_4x10_{p}_{period}')
            model.Add(sum(hrs) + int(l_hrs) - target == gap)
            abs_gap = model.NewIntVar(0, 1000, f'abs_gap_4x10_{p}_{period}')
            model.AddAbsEquality(abs_gap, gap)
            penalty_vars.append(abs_gap * 10000)

    for p in seven_on_providers:
        for period in range(4):
            start_d, end_d = period * 14, (period * 14) + 14
            hrs = [schedule_vars[(p, d, s)] * all_shifts[s][1] for d in range(start_d, end_d) for s in shift_names]
            l_hrs = sum(l['hours'] for l in leave_dict.get(p, []) if l['start'] <= date_list[end_d-1] and l['end'] >= date_list[start_d])

            gap = model.NewIntVar(-1000, 1000, f'gap_{p}_{period}')
            model.Add(sum(hrs) + int(l_hrs) - 80 == gap)
            abs_gap = model.NewIntVar(0, 1000, f'abs_gap_{p}_{period}')
            model.AddAbsEquality(abs_gap, gap)
            penalty_vars.append(abs_gap * 500)

    for p in fee_providers:
        for d in range(num_days):
            for s in shift_names:
                penalty_vars.append(schedule_vars[(p, d, s)] * 800)

    for d in range(num_days):
        if d == 0: continue 
        is_weekend = date_list[d].weekday() >= 5
        is_holiday = date_list[d] in holidays

        for p in providers:
            if p not in fee_providers:
                model.AddAtMostOne([schedule_vars[(p, d, s)] for s in shift_names])
            else:
                model.Add(sum([schedule_vars[(p, d, s)] for s in shift_names]) <= 2)

        for floor_type, s_list in [("morn", [5,6]), ("noon", [12]), ("night", [20])]:
            d_var = model.NewBoolVar(f'd_{floor_type}_{d}')
            model.Add(sum(schedule_vars[(p, d, s)] for p in providers if roles.get(p)=='MD' for s in shift_names if all_shifts[s][0] in s_list) + d_var >= 1)
            penalty_vars.append(d_var * 1000)

        model.Add(sum(schedule_vars[(p, d, s)] for p in providers if roles.get(p) == 'MD' for s in shift_names if all_shifts[s][0] >= 18) <= 2)

        d_night_shortage = model.NewIntVar(0, 3, f'd_night_shortage_{d}')
        model.Add(sum(schedule_vars[(p, d, s)] for p in providers for s in shift_names if all_shifts[s][0] in [14, 16, 18, 20]) + d_night_shortage >= 3)
        penalty_vars.append(d_night_shortage * 1000)

        if not is_weekend and not is_holiday:
            d_pit = model.NewBoolVar(f'd_pit_{d}')
            d_bravomd = model.NewBoolVar(f'd_bravomd_{d}')
            d_bravoapp = model.NewBoolVar(f'd_bravoapp_{d}')

            model.Add(sum(schedule_vars[(p, d, '9a-7p')] for p in providers if roles.get(p) == 'MD') + d_bravomd >= 1)
            model.Add(sum(schedule_vars[(p, d, '10a-8p')] for p in providers if roles.get(p) == 'APP') + d_pit + d_bravoapp >= 2)

            penalty_vars.extend([d_bravomd * 1000, d_pit * 1000, d_bravoapp * 1000])

    model.Minimize(sum(penalty_vars))
    
    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = 8
    solver.parameters.max_time_in_seconds = 1440.0
    status = solver.Solve(model)

    final_schedule = {}
    if status == cp_model.OPTIMAL or status == cp_model.FEASIBLE:
        for p in providers:
            for d in range(num_days):
                assigned_shift = None
                for s in shift_names:
                    if solver.Value(schedule_vars[(p, d, s)]) == 1:
                        assigned_shift = s
                        break
                final_schedule[(p, d)] = assigned_shift
        return final_schedule, status, date_list, all_shifts, shift_names, holidays
    else:
        return None, status, None, None, None, None

def run_block_5(final_schedule, providers, roles, num_days, date_list, holidays, all_shifts):
    unfilled_log = {d: [] for d in range(num_days)}
    for d in range(num_days):
        is_weekend = date_list[d].weekday() >= 5
        is_holiday = date_list[d] in holidays

        daily_shifts = {p: final_schedule[(p, d)] for p in providers if final_schedule[(p, d)] is not None}
        md_shifts = {p: s for p, s in daily_shifts.items() if roles.get(p) == 'MD'}
        app_shifts = {p: s for p, s in daily_shifts.items() if roles.get(p) == 'APP'}

        if not any(s in ['5a-3p', '5a-5p', '6a-4p'] for s in md_shifts.values()): unfilled_log[d].append("Morn AIC")
        if not any(s == '12p-10p' for s in md_shifts.values()): unfilled_log[d].append("Noon AIC")
        if not any(s in ['8p-6a', '8p-8a'] for s in md_shifts.values()): unfilled_log[d].append("Night AIC")

        if not is_weekend and not is_holiday:
            if not any(s == '9a-7p' for s in md_shifts.values()): unfilled_log[d].append("Bravo MD")
            app_10a_count = sum(1 for s in app_shifts.values() if s == '10a-8p')
            if app_10a_count < 1: unfilled_log[d].append("PIT APP")
            if app_10a_count < 2: unfilled_log[d].append("Bravo APP")

        night_shifts = ['8p-6a', '8p-8a', '6p-6a', '6p-4a', '4p-2a', '2p-12a']
        night_workers = [p for p, s in daily_shifts.items() if s in night_shifts]
        if len(night_workers) < 3: unfilled_log[d].append("Night Block (Understaffed)")
    return unfilled_log

def run_block_6(final_schedule, providers, roles, num_days, date_list, holidays, leave_dict, all_shifts, unfilled_log):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PULSE Schedule"
    ws.freeze_panes = "C2"

    shift_colors = {
        '5a-3p': 'FFF2CC', '5a-5p': 'FFF2CC', '6a-4p': 'FFF2CC', '7a-5p': 'FFF2CC', '7a-7p': 'FFF2CC',
        '8a-6p': 'E2EFDA', '9a-7p': 'E2EFDA', '9a-9p': 'E2EFDA', '10a-8p': 'E2EFDA',
        '12p-10p': 'FCE4D6', '2p-12a': 'FCE4D6', '4p-2a': 'F8CBAD',
        '8p-6a': 'B4C6E7', '8p-8a': 'B4C6E7', '6p-6a': 'B4C6E7', '6p-4a': 'B4C6E7'
    }

    red_border = Border(left=Side(style='medium', color='FF0000'), right=Side(style='medium', color='FF0000'), top=Side(style='medium', color='FF0000'), bottom=Side(style='medium', color='FF0000'))
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    ua_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ua_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    headers = ["Initials", "Role"]
    holiday_col_indices, black_col_indices = [], []
    current_col = 3

    for period in range(4):
        start_d, end_d = period * 14, (period * 14) + 14
        for d in range(start_d, end_d):
            headers.append(date_list[d].strftime('%m/%d'))
            if date_list[d] in holidays: holiday_col_indices.append(current_col)
            current_col += 1
        headers.extend(["Worked", "Leave", "Holiday", "Total"])
        current_col += 4
        if period < 3:
            headers.append("SPLIT")
            black_col_indices.append(current_col)
            current_col += 1

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font, cell.alignment = bold_font, center_align

    daily_hours_totals = {d: 0 for d in range(num_days)}

    for p in providers:
        row_data = [p, roles.get(p, "N/A")]
        for period in range(4):
            start_d, end_d = period * 14, (period * 14) + 14
            worked_h, leave_h, holiday_h = 0, 0, 0
            for d in range(start_d, end_d):
                shift = final_schedule.get((p, d))
                is_ua = any(l['start'] <= date_list[d] <= l['end'] for l in leave_dict.get(p, []))
                if shift:
                    row_data.append(shift)
                    worked_h += all_shifts[shift][1]
                    daily_hours_totals[d] += all_shifts[shift][1]
                elif is_ua:
                    row_data.append("UA")
                else:
                    row_data.append("")
                if date_list[d] in holidays: holiday_h += 10
            l_hrs_visual = 0
            for l in leave_dict.get(p, []):
                if l['start'] <= date_list[end_d-1] and l['end'] >= date_list[start_d]:
                    l_hrs_visual += l['hours']
            row_data.extend([worked_h, l_hrs_visual, holiday_h, worked_h + holiday_h + l_hrs_visual])
            if period < 3: row_data.append("")
        ws.append(row_data)

    ws.append([])
    totals_row = ["DAILY TOTAL HOURS", ""]
    for period in range(4):
        start_d, end_d = period * 14, (period * 14) + 14
        for d in range(start_d, end_d): totals_row.append(daily_hours_totals[d])
        totals_row.extend(["", "", "", ""])
        if period < 3: totals_row.append("")
    ws.append(totals_row)
    for col in range(1, len(totals_row) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font, cell.alignment = bold_font, center_align
        if totals_row[col-1] != "": cell.fill = grey_fill

    ws.append([])
    unfilled_row = ["UNFILLED FLOORS", ""]
    for period in range(4):
        start_d, end_d = period * 14, (period * 14) + 14
        for d in range(start_d, end_d):
            unfilled_row.append("--- (UNFILLED): " + ", ".join(unfilled_log[d]) if unfilled_log[d] else "")
        unfilled_row.extend(["", "", "", ""])
        if period < 3: unfilled_row.append("")
    ws.append(unfilled_row)
    for col in range(1, len(unfilled_row) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(color="FF0000", bold=True)

    for row in range(2, ws.max_row - 2):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value in shift_colors:
                cell.fill = PatternFill(start_color=shift_colors[cell.value], end_color=shift_colors[cell.value], fill_type="solid")
            elif cell.value == "UA":
                cell.fill = ua_fill
                cell.font = ua_font
            if col in holiday_col_indices: cell.border = red_border
            if col in black_col_indices: cell.fill = black_fill
            if headers[col-1] == "Total":
                worked_val = ws.cell(row=row, column=col-3).value
                hol_val = ws.cell(row=row, column=col-1).value
                if cell.value == 90 and hol_val >= 10 and worked_val == 80:
                    cell.fill, cell.font = yellow_fill, bold_font

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_length = max((len(str(cell.value)) for cell in col if cell.value and "UNFILLED" not in str(cell.value)), default=0)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 25)
    for col_idx in black_col_indices: ws.column_dimensions[get_column_letter(col_idx)].width = 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# UI: EXECUTION
# ==========================================
if st.button("Generate Pulse Schedule", type="primary"):
    if not roster_file:
        st.warning("Please upload the Roster Excel file first.")
    else:
        with st.spinner("Spinning up PULSE Engine. Crunching 10 million possibilities (Max 24 mins)..."):
            
            dt_prior = pd.to_datetime(anchor_date)
            df_prior, providers, roles = run_block_1(roster_file)
            leave_dict = run_block_2(leave_file, providers)
            
            final_schedule, status, date_list, all_shifts, shift_names, holidays = run_block_3_and_4(
                dt_prior, providers, roles, df_prior, leave_dict
            )
            
            if final_schedule:
                st.success("OPTIMAL Schedule Generated!")
                unfilled_log = run_block_5(final_schedule, providers, roles, 56, date_list, holidays, all_shifts)
                excel_buffer = run_block_6(final_schedule, providers, roles, 56, date_list, holidays, leave_dict, all_shifts, unfilled_log)
                
                st.download_button(
                    label="📥 Download Final Schedule (Excel)",
                    data=excel_buffer,
                    file_name=f"PULSE_Final_{anchor_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("Engine Status: INFEASIBLE. The rules contradict each other; no schedule is mathematically possible.")


